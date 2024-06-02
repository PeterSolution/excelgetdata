import openpyxl
from tkinter import filedialog
from tkinter import Tk
import wordlenght
import math
class excelgetdata:
    def getdata(self):
        try:
            import openpyxl
        except ImportError:
            print("Biblioteka openpyxl nie jest dostępna. Próbuję zainstalować.")
            try:
                import subprocess
                import sys
                subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
                import openpyxl  # Spróbuj ponownie zaimportować po instalacji
            except Exception as e:
                print(f"Błąd instalacji openpyxl: {e}")

        try:
            from tkinter import filedialog, Tk
        except ImportError:
            print("Moduł tkinter nie jest dostępny. Próbuję zainstalować.")
            try:
                import subprocess
                import sys
                subprocess.check_call([sys.executable, "-m", "pip", "install", "tkinter"])
                from tkinter import filedialog, Tk  # Spróbuj ponownie zaimportować po instalacji
            except Exception as e:
                print(f"Błąd instalacji tkinter: {e}")

        try:
            import wordlenght
        except ImportError:
            print("Moduł wordlenght nie jest dostępny. Próbuję zainstalować.")
            try:
                import subprocess
                import sys
                subprocess.check_call([sys.executable, "-m", "pip", "install", "wordlenght"])
                import wordlenght  # Spróbuj ponownie zaimportować po instalacji
            except Exception as e:
                print(f"Błąd instalacji wordlenght: {e}")

        try:
            import math
        except ImportError:
            print("Moduł math nie jest dostępny. Prawdopodobnie jest dostępny w standardowej instalacji Pythona.")

        try:
            import wordfunc
        except ImportError:
            print("Moduł wordfunc nie jest dostępny. Próbuję zainstalować.")
            try:
                import subprocess
                import sys
                subprocess.check_call([sys.executable, "-m", "pip", "install", "wordfunc"])
                import wordfunc  # Spróbuj ponownie zaimportować po instalacji
            except Exception as e:
                print(f"Błąd instalacji wordfunc: {e}")



        root = Tk()
        root.withdraw()

        file_path = filedialog.askopenfilename(
            title="select file",
            filetypes=[("Excel", "*.xlsx"), ("Word", "*.docx"), ("all files", "*.*")]
        )
        excelheight=0
        excelwidth=0
        truedatah=0
        truedatahw=0

        checkclass=wordlenght.check()

        checkifemplty=0
        checkifempltywidth=0
        checkwidth=0
        checkheight=0
        forwidth=10000
        forheight=10000

        maxlenght=0

        if file_path:
            print("file:", file_path)
            sstr = file_path.rfind('/')
            sub = file_path[sstr + 1:-1]

            excel = openpyxl.load_workbook(file_path)

            print("what we have:")
            for sheet_name in excel.sheetnames:
                print(sheet_name)

            try:
                name=excel.sheetnames
                for data in name:
                    data = excel[data]
                    for i in range(forwidth):

                        for j in range(forheight):
                            value = data.cell(row=i + 1, column=j + 1).value
                            if value is not None:

                                checklenght=checkclass.lenght(str(value))

                                if checklenght>maxlenght:
                                    maxlenght=checklenght
                                if excelwidth < j:
                                    excelwidth = j+1
                                if excelheight < i:
                                    excelheight = i+1
                                checkifemplty = 0
                                if (forheight - j) < 10:
                                    forheight = forheight + 1
                            if value is None:
                                checkifemplty = checkifemplty + 1
                                if j < 11 & checkifemplty == 10:
                                    checkifempltywidth = checkifempltywidth + 1
                                else:
                                    if j > 11 & checkifemplty == 10:
                                        break


            finally:
                excel.close()


        print("height: " + str(excelheight))
        print("width: " + str(excelwidth))
        print(str(maxlenght))

        tab = [[0 for _ in range(excelwidth+2)] for _ in range((excelheight*2))]

        for i in range(maxlenght):
            print("_",end="")
        print("")
        if file_path:
            print("file:", file_path)
            sstr = file_path.rfind('/')
            sub = file_path[sstr + 1:-1]

            excel = openpyxl.load_workbook(file_path)

            print("what we have:")
            for sheet_name in excel.sheetnames:
                print(sheet_name)

            try:
                name=excel.sheetnames
                for data in name:
                    data = excel[data]
                    for i in range((excelheight*2)):

                        for j in range(excelwidth+2):

                            value = data.cell(row=i + 1, column=j + 1).value
                            if i == 0 | i == excelwidth+1:
                                print("_",end="")
                            else:
                                if value is not None:
                                    print(f'value from cell ({i + 1},{j + 1}): {value}')
                            if value is not None:
                                tab[i][j-1]=value

            except:
                print("error")

        return tab

